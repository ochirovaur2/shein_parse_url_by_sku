# encoding: utf-8
import requests
import json
import time
import datetime
import re
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import xlrd






input_dir = 'input_xls'
output_file_name = 'output.xlsx'

limit = 1






def get_details(sku):

	url = f'https://ru.shein.com/pdsearch/{sku}'

		
	r = requests.get(url, timeout = 10)
	
	if r.status_code != 200:
		time.sleep(10)
		r = requests.get(url, timeout = 10)

	return (r.status_code, r.text)
		
	
		


def sheet_find_column_indexes(sheet):
	col_indexes = {}
	for col_i in range(sheet.ncols):
		xl_data=sheet.cell(0,col_i).value.replace('\n',' ').strip()
		if 'sku' in xl_data.lower(): 			col_indexes['sku'] = col_i
		
	return col_indexes


def sheet_extract_data(sheet):
	col_indexes = sheet_find_column_indexes(sheet)
	#print(col_indexes)
	#print(sheet.nrows)
	skus = []
	for row_i in range(1, sheet.nrows):
		skus.append( sheet.cell(row_i, col_indexes['sku']).value) 	

	return skus

def get_files_from_dir(output_file_name, dir, exts = ['xls','xlsx']):
	all_files = []
	for root, dirs, files in os.walk(dir):
		for file in files:
			if file.split('.')[-1] in exts:
				all_files.append( os.path.join(root, file))
				output_file_name = file
	return all_files, output_file_name



def write_cell(value, sheet, row_i, col_i, align, font):
	cell = sheet.cell(row=row_i, column=col_i)
	cell.alignment = align
	cell.font = font
	cell.value = value

sendings_dir = os.path.join(os.getcwd(), input_dir)
input_files, output_file_name = get_files_from_dir(output_file_name,sendings_dir, exts=['xls','xlsx'])
print(input_files)
print(f'Input files count: {len(input_files)}')

###################################################################  Get data from xls
i = 0 

for file_path in input_files:
	i += 1
	book = xlrd.open_workbook(file_path)
	sheet = book.sheet_by_index(0)
	skus = sheet_extract_data(sheet)
	
	print(f'File_path\t{file_path}\torders cnt: {len(skus)}')
	
print(f'All orders count: {len(skus)}')





#################################################################### Split list
# split sheet_orders
# How many elements each 
# list should have 




# #################################################################### Save data to xls
book = Workbook()
sheet = book.active
align_default = Alignment(horizontal='left', vertical='center')
align_center =  Alignment(horizontal='center', vertical='center')
font = Font(size = "10", name = 'Arial')

################### заголовок 
write_cell('SKU', 		sheet, 1, 1, align_center, font)
write_cell('URL', 		sheet, 1, 2, align_center, font)
write_cell('Name', 		sheet, 1, 3, align_center, font)

for id, sku in enumerate(skus, start=1):
	
	row_i = id + 1
	try:
		
		print(f'{id}/{len(skus)}')

		
		
		data = get_details(sku)
		
		

		url = json.loads('{'+re.search(r'"goodsDetailUrl":"[a-zA-Z0-9_.+-:#]+"', data[1]).group(0)+'}')['goodsDetailUrl']
		goods_name = json.loads('{'+re.search(r'"goods_name":"(.*?)"', data[1]).group(0)+'}')['goods_name']
		full_url = f"https://ru.shein.com{url}"
		

		
		print(f'{sku}: {full_url}')
		write_cell(sku, 		sheet, row_i, 1, align_default, font)
		write_cell(full_url, 		sheet, row_i, 2, align_default, font)
		write_cell(goods_name, 		sheet, row_i, 3, align_default, font)
		
	
	except:
		print('ошибка', sku)
		write_cell(sku, 		sheet, row_i, 1, align_default, font)
		write_cell("ошибка", 		sheet, row_i, 2, align_default, font)
		write_cell("ошибка", 		sheet, row_i, 3, align_default, font)
	
	
		
book.save(output_file_name)

